#!/usr/bin/env python3
"""Generate Bluesky OMS master data Excel templates."""

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path('/Users/jackiets/Downloads/Copy of Order At Ease (COM) Template.xlsx')
ASSETS = ROOT / 'public' / 'assets'

HEADER_FILL = PatternFill('solid', fgColor='DEE0BB')
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')

PRODUCT_HEADERS = [
    'UOM', 'Product Category', 'Name', 'Description', 'SKU', 'Price', 'Weight',
    'Images', 'Status', 'Remarks', 'NOS', 'Show Weight', 'Show Quantity',
    'Sell In', 'Options',
]

PRODUCT_ROWS = [
    [
        'kg', 'Fresh Fish', 'Tilapia', 'Fresh tilapia; bill by weight per qty',
        'TIL-001', 30.00, 1.0, 'tilapia.jpg', 'active', '', 1, 'No', 'Yes',
        'qty_bill_weight', '{"Situation":["Whole","Cleaned","Fillet"]}',
    ],
    [
        'kg', 'Shellfish', 'Tiger Prawn', 'Sold by weight',
        'PRAWN-001', 45.00, 1.0, 'tiger-prawn.jpg', 'active', '', 1, 'No', 'Yes',
        'weight', '{"Size":["M","L","XL"]}',
    ],
    [
        'kg', 'Fresh Fish', 'Red Snapper', 'Sold by weight',
        'SNAP-001', 55.00, 1.0, 'red-snapper.jpg', 'active', '', 1, 'No', 'Yes',
        'weight', '',
    ],
    [
        'pcs', 'Supplies', 'Ice Box', 'Sold by quantity',
        'ICE-001', 5.00, 0.5, '', 'active', '', 1, 'Yes', 'Yes',
        'qty', '',
    ],
]

CUSTOMER_HEADERS = [
    'Customer Name', 'SQL Customer Code', 'Customer Email', 'Category',
    'Attn. Name', 'Attn. Contact', 'Product Price Permission',
    'Invoice Visibility', 'Invoice Product Price Visibility', 'Area',
    'Billing Address', 'Billing City', 'Billing Postcode', 'Billing State',
    'Shipping Address', 'Shipping City', 'Shipping Postcode', 'Shipping State',
    'Payment Method', 'Remark', 'Password',
]

CUSTOMER_ROWS = [
    [
        'Ocean Bistro', 'BS-CUST-001', 'order@oceanbistro.com', 'Restaurant',
        'Mr Tan', '0123456789', 'Yes', 'Yes', 'Yes', 'George Town',
        '12 Jalan Macalister', 'George Town', '10400', 'Penang',
        '12 Jalan Macalister', 'George Town', '10400', 'Penang',
        'cod,bank-transfer', 'COD preferred before 6pm', 'ChangeMe123!',
    ],
    [
        'Harbour Seafood', 'BS-CUST-002', 'accounts@harbourseafood.com', 'Wholesale',
        'Ms Lim', '0167788990', 'Yes', 'Yes', 'No', 'Butterworth',
        '88 Jalan Chain Ferry', 'Butterworth', '12000', 'Penang',
        '88 Jalan Chain Ferry', 'Butterworth', '12000', 'Penang',
        'term,bank-transfer', 'Credit term customer', 'ChangeMe123!',
    ],
    [
        'Walk-in Counter', 'BS-CASH-001', '', 'Retail',
        'Counter', '0112233445', 'No', 'No', 'No', 'George Town',
        '1 Jalan Transfer', 'George Town', '10050', 'Penang',
        '', '', '', '',
        'cod', 'Walk-in; no email login', 'ChangeMe123!',
    ],
]

TEXT_COLUMNS = {
    'Customers': {'B', 'F', 'M', 'Q', 'U'},
}

README_LINES = [
    'Bluesky Live Seafood OMS — Master Data Import Template',
    '',
    'Use the Products and Customers sheets for import. Other sheets are reference or report layouts.',
    '',
    'Products import: Admin → Products → Import. Use column order exactly as shown in the Products sheet.',
    'Customers import: Admin → Customers → Import. Use column order exactly as shown in the Customers sheet.',
    '',
    'Sell In values: qty | weight | qty_bill_weight',
    'Product status: active | inactive',
    'Payment methods (comma-separated): cod, term, bank-transfer, e-wallet, cash, qr, credit-term, customer-credit',
    'Yes/No fields: Product Price Permission, Invoice Visibility, Invoice Product Price Visibility, Show Weight, Show Quantity',
    '',
    'Notes:',
    '- UOM and Product Category are auto-created during product import if they do not exist.',
    '- Images should be filenames already uploaded to the system, or a JSON array e.g. ["a.jpg","b.jpg"].',
    '- Options must be valid JSON, e.g. {"Situation":["Whole","Cleaned"]}.',
    '- Keep SQL Customer Code, phone, and postcodes as text to avoid Excel scientific notation.',
]

REFERENCE_SELL_IN = [
    ['Sell In', 'Description', 'Show Qty', 'Show Weight', 'Example'],
    ['qty', 'Bill and deduct by quantity', 'Yes', 'Optional', 'Ice box, packaging'],
    ['weight', 'Bill and deduct by weight (kg)', 'No', 'Yes', 'Prawn, fish by kg'],
    ['qty_bill_weight', 'Customer orders qty; price from weight x unit price', 'Yes', 'Yes', 'Tilapia: 1 fish, 3kg, bill RM90'],
]

REFERENCE_PAYMENTS = [
    ['Code', 'Description', 'Typical use'],
    ['cod', 'Cash on delivery', 'Retail / walk-in delivery'],
    ['term', 'Payment term / credit account', 'Regular wholesale customers'],
    ['bank-transfer', 'Bank transfer', 'Invoiced customers'],
    ['e-wallet', 'E-wallet', 'Online payment'],
    ['cash', 'Cash', 'Counter payment'],
    ['qr', 'QR payment', 'Scan & pay'],
    ['credit-term', 'Credit term (legacy alias)', 'Same as term'],
    ['customer-credit', 'Customer credit balance', 'Uses stored credit balance'],
]

REFERENCE_DRIVERS = [
    ['Name', 'Username', 'Phone', 'Status', 'Notes'],
    ['Ahmad Rizal', 'driver1', '0123001001', 'Active', 'Create via Admin → Drivers (no Excel import yet)'],
    ['Kumar Raj', 'driver3', '0123001003', 'Active', ''],
    ['Tan Wei Ming', 'driver2', '0123001002', 'Active', ''],
]

REPORT_SHEETS = {
    'Daily Sales Report': [
        'No', 'Order At', 'Customer', 'Item Name', 'SKU', 'Category', 'Quantity',
        'Unit Price', 'Total Price', 'Payment Method', 'Area', 'Billing Address',
        'Shipping Address', 'Last Updated At',
    ],
    'Daily Summary Report': [
        'No.', 'Order At', 'Customer', 'Item Name', 'SKU', 'Category', 'Quantity',
        'Unit Weight', 'Total Weight',
    ],
    'Daily Summary Stock': [
        'Item Name', 'Item SKU', 'Item Category', 'Item Quantity',
    ],
    'DO Report': [
        'Order ID', 'Order At', 'Customer', 'Weight', 'Products', 'Total Price',
        'Payment Method', 'Area', 'Billing Address', 'Shipping Address', 'Driver',
        'Fulfillment', 'Status', 'Last Updated At',
    ],
    'SQL DO Export': [
        'Order ID', 'Customer', 'Area', 'Billing Address', 'Shipping Address',
        'Driver', 'Fulfillment', 'Status', 'Last Updated At',
    ],
    'UOM Reference': ['UOM', 'Notes', 'Created via product import'],
    'Category Reference': ['Category', 'Notes', 'Created via product import'],
}


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP


def autosize_columns(ws, max_width=42):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), max_width))
        ws.column_dimensions[letter].width = max(12, max_len + 2)


def write_table(ws, headers, rows, text_cols=None):
    ws.append(headers)
    style_header_row(ws)
    for row in rows:
        ws.append(row)
    if text_cols:
        for r in range(2, ws.max_row + 1):
            for col in text_cols:
                cell = ws[f'{col}{r}']
                cell.number_format = '@'
                if cell.value is not None:
                    cell.value = str(cell.value)
    autosize_columns(ws)


def write_readme(ws):
    ws.title = 'README'
    ws['A1'] = README_LINES[0]
    ws['A1'].font = Font(bold=True, size=14)
    row = 3
    for line in README_LINES[2:]:
        ws.cell(row=row, column=1, value=line)
        ws.cell(row=row, column=1).alignment = WRAP
        row += 1
    ws.column_dimensions['A'].width = 110


def build_master_workbook():
    wb = Workbook()
    write_readme(wb.active)

    ws_products = wb.create_sheet('Products')
    write_table(ws_products, PRODUCT_HEADERS, PRODUCT_ROWS)

    ws_customers = wb.create_sheet('Customers')
    write_table(ws_customers, CUSTOMER_HEADERS, CUSTOMER_ROWS, TEXT_COLUMNS['Customers'])

    ws_sell = wb.create_sheet('Reference - Sell In')
    write_table(ws_sell, REFERENCE_SELL_IN[0], REFERENCE_SELL_IN[1:])

    ws_pay = wb.create_sheet('Reference - Payments')
    write_table(ws_pay, REFERENCE_PAYMENTS[0], REFERENCE_PAYMENTS[1:])

    ws_drivers = wb.create_sheet('Drivers (Reference)')
    write_table(ws_drivers, REFERENCE_DRIVERS[0], REFERENCE_DRIVERS[1:])

    for title, headers in REPORT_SHEETS.items():
        ws = wb.create_sheet(title)
        write_table(ws, headers, [])

    return wb


def build_single_sheet_workbook(headers, rows, text_cols=None, title='Sheet1'):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    write_table(ws, headers, rows, text_cols)
    return wb


def main():
    ASSETS.mkdir(parents=True, exist_ok=True)

    master = build_master_workbook()
    master.save(DOWNLOADS)

    products = build_single_sheet_workbook(PRODUCT_HEADERS, PRODUCT_ROWS, title='Products')
    products.save(ASSETS / 'product_template.xlsx')

    customers = build_single_sheet_workbook(
        CUSTOMER_HEADERS,
        CUSTOMER_ROWS,
        TEXT_COLUMNS['Customers'],
        title='Customers',
    )
    customers.save(ASSETS / 'customers_sample.xlsx')

    print('Wrote:')
    print(' ', DOWNLOADS)
    print(' ', ASSETS / 'product_template.xlsx')
    print(' ', ASSETS / 'customers_sample.xlsx')


if __name__ == '__main__':
    main()
