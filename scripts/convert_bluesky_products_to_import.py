#!/usr/bin/env python3
"""Convert BLUESKY PRODUCT NAME.xlsx to the OMS admin product import template."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print('Missing dependency: pip install openpyxl', file=sys.stderr)
    raise SystemExit(1)

IMPORT_HEADERS = [
    'UOM', 'Product Category', 'Name', 'Description', 'SKU', 'Price', 'Weight',
    'Images', 'Status', 'Remarks', 'NOS', 'Show Weight', 'Show Quantity',
    'Sell In', 'Options',
]


def clean(value) -> str:
    return re.sub(r'\s+', ' ', str(value or '').strip())


def sell_in_for_uom(uom: str) -> tuple[str, str, str]:
    uom = uom.upper()
    if uom in {'PCS', 'BAG', 'BOX', 'PKT', 'PACK'}:
        return 'qty', 'No', 'Yes'
    return 'weight', 'Yes', 'No'


def parse_source(path: Path) -> list[list]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    products = []
    seen = set()

    for index, row in enumerate(rows):
        if index < 2:
            continue

        cells = list(row) + [''] * (5 - len(row))
        uom, sku, category, name, chinese = [clean(c) for c in cells[:5]]

        if not sku or not name:
            continue
        if sku.upper() == 'PRODUCT CODE':
            continue

        sku = sku.upper()
        if sku in seen:
            continue
        seen.add(sku)

        uom = (uom or 'KG').upper()
        sell_in, show_weight, show_qty = sell_in_for_uom(uom)
        description = f'{name} / {chinese}' if chinese else name

        products.append([
            uom,
            category or 'UNCATEGORIZED',
            name,
            description[:200],
            sku,
            0,
            1,
            '',
            'active',
            '',
            1,
            show_weight,
            show_qty,
            sell_in,
            '',
        ])

    return products


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        'source',
        nargs='?',
        default='/Users/jackiets/Downloads/BLUESKY PRODUCT NAME.xlsx',
        help='Path to BLUESKY PRODUCT NAME.xlsx',
    )
    parser.add_argument(
        '-o',
        '--output',
        default='storage/app/bluesky_products_import.xlsx',
        help='Output path for OMS import template',
    )
    args = parser.parse_args()

    source = Path(args.source).expanduser()
    output = Path(args.output)
    if not output.is_absolute():
        output = Path(__file__).resolve().parents[1] / output

    if not source.exists():
        print(f'Source file not found: {source}', file=sys.stderr)
        return 1

    products = parse_source(source)
    if not products:
        print('No product rows found.', file=sys.stderr)
        return 1

    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    ws.append(IMPORT_HEADERS)
    for row in products:
        ws.append(row)
    wb.save(output)

    print(f'Wrote {len(products)} product(s) to {output}')
    print('Upload via Admin → Products → Import Products')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
